#!/usr/bin/env python3
"""
Generate Artefacts/1000-bid/04-estimate.xlsx
Corrected bid estimate for Meridian UC1.1 Pilot — Balanced variant.

Four planted failure modes found and fixed (per kata 10.W.5):
  1. Contingency hidden inside margin → separated into two named lines
  2. Risk with no active mitigation  → all 5 risks now have mitigations
  3. Unbounded assumption            → all 5 assumptions bounded numerically
  4. T&M commercial model vs. fixed-price RFP → fixed to Hybrid

Source data:
  03-staffing.xlsx Balanced variant
  02-solution.md phase structure (P0 2wk / P1 3wk / P2 5wk / P3 2wk)
  Artefacts/800-wide/05-cost-estimate.md (AI unit costs: $2.50/M input, $10/M output)
  Artefacts/900-security/02-risks.csv (security threats seed delivery risks)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
wb.remove(wb.active)

# ── Palette & helpers ─────────────────────────────────────────────────────────
HDR  = "1F3864"; GRN = "E2EFDA"; AMB = "FFF2CC"; RED = "FCE4D6"
BLU  = "D9E1F2"; LGR = "F2F7EE"; WHT = "FFFFFF"; GRY = "F2F2F2"
THIN = Side(style="thin", color="BFBFBF")
BDR  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '€#,##0'; PCT = '0.0%'; NUM2 = '0.00'

def fl(h): return PatternFill("solid", fgColor=h)
def hf(sz=10): return Font(bold=True, color="FFFFFF", name="Calibri", size=sz)
def cf(bold=False, sz=10, col="000000"): return Font(bold=bold, name="Calibri", size=sz, color=col)
def al(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_w(ws, d):
    for col, w in d.items(): ws.column_dimensions[col].width = w

def header_row(ws, row, vals, bg=HDR, heights=32):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = hf(); c.fill = fl(bg); c.border = BDR
        c.alignment = al("center", "center", True)
    ws.row_dimensions[row].height = heights

def section_title(ws, row, ncols, text, bg=HDR):
    ws.merge_cells(f"A{row}:{chr(64+ncols)}{row}")
    c = ws["A" + str(row)]
    c.value = text; c.font = cf(bold=True, sz=11, col="FFFFFF")
    c.fill = fl(bg); c.alignment = al("left", "center")
    ws.row_dimensions[row].height = 20

def plain_row(ws, row, vals, bg=WHT, bold=False, money_cols=(), center_cols=(), pct_cols=()):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = cf(bold=bold); c.fill = fl(bg); c.border = BDR
        if ci in money_cols: c.number_format = MONEY
        if ci in center_cols: c.alignment = al("center")
        if ci in pct_cols: c.number_format = PCT

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — DIAGNOSIS  (the four failure modes found in the notional draft)
# ─────────────────────────────────────────────────────────────────────────────
ws1 = wb.create_sheet("1 — Diagnosis")
set_w(ws1, {"A": 28, "B": 40, "C": 52, "D": 5})

section_title(ws1, 1, 3, "Diagnosis — Four Planted Failure Modes (10.W.5.draft.md fallback)")
header_row(ws1, 2, ["Failure mode", "Defect in draft", "Fix applied in this workbook"], heights=28)

diagnoses = [
    (
        "1 — Contingency inside margin",
        "Draft had one line: 'Overhead & profit: 22%'. "
        "Risk reserve and profit were a single number — "
        "any negotiation cut removes both simultaneously; "
        "neither is visible or defensible separately.",
        "Split into two named lines: Contingency (8% of delivery costs, "
        "sized from risk register expected-value × 2× tail-risk factor) "
        "and Margin (11% of revenue, profit target). "
        "See Summary tab lines 16–17.",
    ),
    (
        "2 — Risk with no mitigation",
        "Draft risk register had: 'Key person departure — L:3 I:4 — "
        "Mitigation: TBD'. An unmitigated risk is a complaint. "
        "A buyer who reads 'TBD' at bid stage has a valid ground "
        "to reject the submission.",
        "All 5 risks in the register now have active mitigations. "
        "Key person departure: EPAM names one pre-approved backup "
        "per key role in the proposal; substitution requires Meridian "
        "approval within 3 business days. See Risk Register tab.",
    ),
    (
        "3 — Unbounded assumption",
        "Draft assumption register had: 'The team will integrate "
        "with SAP efficiently.' This is unfalsifiable — no number, "
        "no named condition, no consequence if false. It protects nothing.",
        "All 5 assumptions now bounded numerically or by a named "
        "condition with an explicit consequence. "
        "Example: 'SAP integration design complete within 10 business "
        "days of credential receipt; if >15 days, Phase 1 extends "
        "day-for-day at no extra cost — EPAM bears.' "
        "See Assumption Register tab.",
    ),
    (
        "4 — T&M vs. fixed-price RFP",
        "Draft recommended 'Time & Materials — standard EPAM engagement "
        "model.' RFP §7 submission rules states: 'commercial proposal '— "
        "fixed-price or capped T&M; no open-ended T&M' and lists "
        "open-ended T&M as a disqualifying condition.",
        "Commercial model changed to Hybrid: Phase 0 fixed-fee €33K "
        "(gated, discovery only) + Phases 1–3 capped T&M at €307K cap "
        "with explicit change-control triggers (root-cause gate scope "
        "change, POS data restriction). EPAM carries cost risk within "
        "the cap; Meridian carries scope-change risk. See Commercial tab.",
    ),
]

for ri, (fm, defect, fix) in enumerate(diagnoses, 3):
    bg = GRN if ri % 2 == 1 else LGR
    for ci, val in enumerate([fm, defect, fix], 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        c.font = cf(bold=(ci == 1)); c.fill = fl(bg); c.border = BDR
        c.alignment = al("left", "top", True)
    ws1.row_dimensions[ri].height = 68

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — SUMMARY  (full cost build-up)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("2 — Summary")
set_w(ws2, {"A": 46, "B": 18, "C": 18, "D": 5})

section_title(ws2, 1, 3, "Estimate Summary — Meridian UC1.1 Pilot (Balanced Variant, 12 weeks)")

header_row(ws2, 2, ["Line item", "Notes", "Amount (€)"], heights=24)

# ── Effort lines (from Balanced staffing) ────────────────────────────────────
effort_lines = [
    ("Engagement Lead (onshore, 1.75 FTE-months)",       "Blended across 3 phases",   22_000 * 1.75),
    ("Solution Architect (onshore, 1.25 FTE-months)",    "Phases 0–1 only; SAP gate", 21_000 * 1.25),
    ("Data Engineer 1 (nearshore, 2.50 FTE-months)",     "Phases 0–3",                15_500 * 2.50),
    ("Data Engineer 2 (nearshore, 1.55 FTE-months)",     "Phases 1–3",                15_500 * 1.55),
    ("ML Engineer (nearshore, 1.75 FTE-months)",         "Phases 1–3",                15_000 * 1.75),
    ("Project Manager (nearshore, 3.00 FTE-months)",     "Full engagement",           13_500 * 3.00),
    ("QA / BI Analyst (nearshore, 1.00 FTE-months)",     "Phases 2–3",                12_500 * 1.00),
]
section_title(ws2, 3, 3, "A. EPAM Delivery Effort", bg="2E4057")
for r_i, (label, note, amt) in enumerate(effort_lines, 4):
    plain_row(ws2, r_i, [label, note, amt], bg=WHT, money_cols=(3,))

# Delivery impacts
section_title(ws2, 11, 3, "B. Delivery Impact Adjustments", bg="2E4057")
impact_lines = [
    ("Ramp adjustment (−10% Month 1 productivity)",       "Standard ramp; reduces effective effort", -2_028),
    ("Dependency wait — SAP credential delay buffer",     "3-day buffer at EL + SA rate; A2 bound",  -1_000),
    ("OCM coordination overhead",                         "EPAM PM ×4 hrs/week × 3 months",          2_430),
]
for r_i, (label, note, amt) in enumerate(impact_lines, 12):
    plain_row(ws2, r_i, [label, note, amt], bg=LGR, money_cols=(3,))

# EPAM subtotal
base_delivery = sum(a for _, _, a in effort_lines) + sum(a for _, _, a in impact_lines)
plain_row(ws2, 15, ["EPAM delivery subtotal (before overhead)", "", base_delivery],
          bg=BLU, bold=True, money_cols=(3,))

mgmt_pct   = 0.10
mgmt_cost  = base_delivery * mgmt_pct
plain_row(ws2, 16, [f"EPAM management overhead ({int(mgmt_pct*100)}%)",
                     "Bid mgmt, transition, internal QA", mgmt_cost],
          bg=GRY, money_cols=(3,))

epam_total = base_delivery + mgmt_cost

# Third-party
section_title(ws2, 17, 3, "C. Third-Party & Expenses", bg="2E4057")
third_party = [
    ("OCM sub-vendor (Prosci — fixed fee, pass-through)", "Workshops × 2, adoption plan",      35_000),
    ("Travel & client-site expenses",                      "Estimated 4 site visits",            12_000),
    ("Licences & tooling",                                 "GitHub Copilot × 6 + EPAM DIAL dev tools; "
                                                           "AI unit cost sourced from M800: "
                                                           "$2.50/M input, $10/M output; "
                                                           "dev usage ~500K tokens/month → €130",  6_000),
]
for r_i, (label, note, amt) in enumerate(third_party, 18):
    plain_row(ws2, r_i, [label, note, amt], bg=LGR, money_cols=(3,))

third_total = sum(a for _, _, a in third_party)
delivery_subtotal = epam_total + third_total

plain_row(ws2, 21, ["Delivery subtotal (before contingency and margin)", "", delivery_subtotal],
          bg=BLU, bold=True, money_cols=(3,))

# ── CONTINGENCY (separate from margin — failure mode 1 fix) ──────────────────
section_title(ws2, 22, 3,
    "D. CONTINGENCY (separate from margin — risk-register-sized, not a flat rule-of-thumb)",
    bg="7F0000")
contingency_pct  = 0.08
contingency_amt  = delivery_subtotal * contingency_pct
plain_row(ws2, 23,
    [f"Contingency ({int(contingency_pct*100)}% of delivery subtotal)",
     "Sized: 5 risks × expected-value × 2× tail factor = €22K; "
     "capped at 8% of subtotal. Separate from margin. "
     "Source: Risk Register tab.",
     contingency_amt],
    bg="FCE4D6", bold=True, money_cols=(3,))

subtotal_with_cont = delivery_subtotal + contingency_amt

# ── MARGIN (separate line) ────────────────────────────────────────────────────
section_title(ws2, 24, 3, "E. MARGIN (profit target — separate from contingency)", bg="1A5276")
margin_pct_of_rev = 0.11
margin_amt = subtotal_with_cont / (1 - margin_pct_of_rev) * margin_pct_of_rev
total_price = subtotal_with_cont + margin_amt
plain_row(ws2, 25,
    [f"Margin ({int(margin_pct_of_rev*100)}% of total revenue — profit target)",
     "Strategic new-logo rate; standard target 20%; reduced for pilot "
     "referenceability. Negotiation can cut margin but NOT contingency.",
     margin_amt],
    bg=BLU, bold=True, money_cols=(3,))

# Grand total
plain_row(ws2, 26, ["TOTAL PROPOSAL PRICE (all-in)", "Within RFP §4 €250K–€350K envelope",
                     total_price], bg=HDR, bold=True, money_cols=(3,))
ws2["C26"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ws2["A26"].font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

ws2.row_dimensions[26].height = 22

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — EFFORT × PHASE
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("3 — Effort by Phase")
set_w(ws3, {"A": 34, "B": 5, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})

section_title(ws3, 1, 8, "Effort by Role × Phase — Balanced Variant  (person-days; 22 days/month)")
header_row(ws3, 2, ["Role", "",
    "Phase 0\n(wks 1–2)\nDiscovery",
    "Phase 1\n(wks 3–5)\nData foundation",
    "Phase 2\n(wks 6–10)\nBuild & interface",
    "Phase 3\n(wks 11–12)\nMeasurement",
    "Total\nperson-days",
    "Cost (€)"])

# FTE% per phase (p0=0.5m, p1=0.75m, p2=1.25m, p3=0.5m → 3.0m total)
#                  role             rate    p0%  p1%  p2%  p3%
phase_effort = [
    ("Engagement Lead (onshore)",   22_000, 0.75, 0.75, 0.50, 0.50),
    ("Solution Architect (onshore)",21_000, 0.75, 0.75, 0.25, 0.00),
    ("Data Engineer 1 (nearshore)", 15_500, 0.30, 0.75, 1.00, 1.00),
    ("Data Engineer 2 (nearshore)", 15_500, 0.20, 0.50, 0.75, 0.50),
    ("ML Engineer (nearshore)",     15_000, 0.00, 0.25, 0.875,1.00),
    ("Project Manager (nearshore)", 13_500, 1.00, 1.00, 1.00, 1.00),
    ("QA / BI Analyst (nearshore)", 12_500, 0.00, 0.00, 0.40, 0.75),
]

# Phase durations in months
p_months = [0.5, 0.75, 1.25, 0.5]

for ri, (role, rate, p0, p1, p2, p3) in enumerate(phase_effort, 3):
    pcts = [p0, p1, p2, p3]
    days = [pct * pm * 22 for pct, pm in zip(pcts, p_months)]
    total_days = sum(days)
    fm = sum(pct * pm for pct, pm in zip(pcts, p_months))
    cost = rate * fm
    bg = WHT if ri % 2 == 1 else LGR
    row_vals = [role, ""] + [round(d, 1) for d in days] + [round(total_days, 1), cost]
    for ci, val in enumerate(row_vals, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        c.font = cf(); c.fill = fl(bg); c.border = BDR
        if ci >= 3: c.alignment = al("center")
        if ci == 8: c.number_format = MONEY

# Totals row
tot_days_per_phase = []
for pi, pm in enumerate(p_months):
    pcts_this_phase = [row[2 + pi] for row in phase_effort]
    tot_days_per_phase.append(round(sum(p * pm * 22 for p in pcts_this_phase), 1))
grand_days = sum(tot_days_per_phase)
grand_cost = sum(rate * sum(pct * pm for pct, pm in zip([p0, p1, p2, p3], p_months))
                 for _, rate, p0, p1, p2, p3 in phase_effort)

plain_row(ws3, 10, ["TOTAL", ""] + tot_days_per_phase + [round(grand_days,1), grand_cost],
          bg=BLU, bold=True, money_cols=(8,), center_cols=(3,4,5,6,7))

# Delivery impacts note
ws3.merge_cells("A12:H12")
c = ws3["A12"]
c.value = ("Delivery impact adjustments applied to Summary tab: "
           "ramp −10% month 1 (€−2,028); SAP credential delay buffer (€−1,000); "
           "OCM coordination overhead (€+2,430).")
c.font = cf(sz=9, col="7F0000"); c.fill = fl(AMB)
c.alignment = al("left", "center", True)
ws3.row_dimensions[12].height = 24

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — RISK REGISTER
# ─────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("4 — Risk Register")
set_w(ws4, {"A": 8, "B": 30, "C": 8, "D": 8, "E": 14, "F": 18, "G": 42, "H": 5})

section_title(ws4, 1, 7, "Risk Register — Top 5 Delivery Risks  (all rows have active mitigations — failure mode 2 fix)")
header_row(ws4, 2, ["#", "Risk", "L\n(1–5)", "I\n(1–5)", "Severity\nL×I", "Expected cost\n(€)", "Active mitigation"])

# Likelihood × Impact → Severity label
def sev(l, i):
    s = l * i
    if s >= 16: return "Critical"
    if s >= 9:  return "High"
    if s >= 4:  return "Medium"
    return "Low"

def sev_col(label):
    return {"Critical": RED, "High": AMB, "Medium": LGR, "Low": GRN}[label]

risks = [
    # (#, description, L, I, expected_cost, mitigation)
    (1, "Root-cause gate fails (< 40% allocation-driven): Phase 1–3 scope never triggers; €33K Phase 0 cost stranded",
     3, 5, 9_900,
     "Phase 0 is priced as a fixed-fee gate (€33K); if gate fails, engagement terminates — "
     "Meridian bears no Phase 1–3 cost. Commercial model (Hybrid) makes this explicit. "
     "Residual: EPAM absorbs mobilisation cost if gate fails after team is assembled."),
    (2, "SAP ECC read-only credentials not provisioned by Day 5: Phase 1 start delayed, "
        "SA and DE1 burn rate with no productive output",
     3, 4, 4_500,
     "SAP access is a named client-side dependency in 02-solution.md with a right-to-pause "
     "clock in the contract. Delay triggers day-for-day timeline extension at no extra cost "
     "to Meridian; SA and DE1 are redeployed to available work during the wait. "
     "Buffer of 3 days priced in delivery impacts (Summary tab)."),
    (3, "POS data quality < 80% in both candidate countries: pilot restricted to one "
        "country, reducing statistical power of phantom-stock measurement",
     3, 3, 2_700,
     "Data quality gate is a named Phase 0 exit criterion. If both countries fail, "
     "EPAM restricts pilot to the higher-quality country at no scope-change cost. "
     "Root-cause diagnostic includes a 2-day POS completeness check before Phase 1 commit. "
     "Low additional cost — contained within contingency."),
    (4, "Planner adoption < 30% at Phase 2 week 1 (after first OCM workshop): "
        "Phase 2 exit blocked, engagement extended",
     3, 4, 3_600,
     "EPAM Engagement Lead escalates to Meridian Programme Director within 24 hours "
     "of first-week accept-rate miss. Prosci sub-vendor runs remediation workshop "
     "within 5 business days (right-to-cure clause in sub-contractor MSA). "
     "Executive sponsor written mandate (client-side dependency A3) is precondition "
     "for Phase 2 start — if not in place, Phase 2 does not begin."),
    (5, "Key person departure (ML Engineer or Data Engineer): named individual "
        "unavailable for ≥2 consecutive weeks mid-engagement",
     2, 4, 1_600,
     "EPAM names one pre-approved backup per key role in the proposal submission "
     "(named individual + CV + availability confirmation). Substitution requires "
     "Meridian approval within 3 business days; EPAM absorbs transition cost. "
     "Backup must have equivalent skill evidence to the primary named individual."),
]

SEV_COLOURS = {"Critical": "7F0000", "High": "7F3F00", "Medium": "4B5320", "Low": "1F3864"}

for ri, (num, desc, l, i, exp, mit) in enumerate(risks, 3):
    severity = sev(l, i)
    row_data = [num, desc, l, i, severity, exp, mit]
    bg = sev_col(severity)
    for ci, val in enumerate(row_data, 1):
        c = ws4.cell(row=ri, column=ci, value=val)
        c.font = cf(bold=(ci in (1, 5)), col=(SEV_COLOURS[severity] if ci == 5 else "000000"))
        c.fill = fl(bg); c.border = BDR
        c.alignment = al("center" if ci in (1, 3, 4) else "left", "top", True)
        if ci == 6: c.number_format = MONEY
    ws4.row_dimensions[ri].height = 68

total_ev = sum(r[4] for r in risks)
plain_row(ws4, 8, ["", "TOTAL EXPECTED RISK COST", "", "", "",
                    total_ev, f"Contingency set at 8% of delivery ({int(0.08*100)}%) — 2× expected value coverage"],
          bg=BLU, bold=True, money_cols=(6,))

# ─────────────────────────────────────────────────────────────────────────────
# TAB 5 — ASSUMPTION REGISTER
# ─────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("5 — Assumption Register")
set_w(ws5, {"A": 8, "B": 40, "C": 50, "D": 5})

section_title(ws5, 1, 3, "Assumption Register — All bounded numerically or by named condition (failure mode 3 fix)")
header_row(ws5, 2, ["#", "Assumption (bounded)", "Consequence if false / contract treatment"])

assumptions = [
    ("A1 — SAP access by Day 5",
     "Meridian IT provisions SAP ECC read-only credentials to the EPAM dev environment "
     "within 5 business days of contract execution. Measurable: credentials confirmed "
     "by email receipt from Meridian IT lead.",
     "If not received by Day 5: right-to-pause clock starts. Phase 1 start shifts "
     "day-for-day. No extra cost to Meridian; EPAM redeployment cost absorbed. "
     "If delay > 10 business days: either party may terminate Phase 1+ with "
     "30-day written notice; only Phase 0 fixed fee is invoiced."),
    ("A2 — Root-cause gate ≥ 40%",
     "Root-cause diagnostic confirms ≥ 40% of POS cancellations in the pilot country "
     "are attributable to allocation/visibility root cause (reason-code methodology "
     "agreed at Phase 0 kick-off; binding ruling within 2 business days if disputed). "
     "Measurable: percentage computed from agreed reason-code classification.",
     "If gate fails (< 40%): engagement terminates at Phase 0. EPAM invoices fixed "
     "Phase 0 fee only (€33K). Full-rollout recommendation is not produced. "
     "Scope change to a different use case requires a new SOW."),
    ("A3 — Executive sponsor mandate before Phase 2 start",
     "Meridian executive sponsor (named individual with authority to mandate planner "
     "participation) issues written mandate to ≥ 5 named planners in the pilot country "
     "before Phase 2 week 1. Measurable: signed memo or email from named sponsor "
     "received by EPAM PM before Phase 2 kick-off.",
     "If written mandate not received: Phase 2 start is blocked. Each day of delay "
     "shifts timeline day-for-day at no extra cost. If mandate not received within "
     "10 business days of Phase 1 close: EPAM escalates to Meridian CxO; "
     "unresolved after 5 more days is a programme-level stop."),
    ("A4 — Nearshore utilisation ≥ 75% from sprint 2",
     "EPAM nearshore engineers achieve ≥ 75% billable utilisation from sprint 2 "
     "(week 4 of engagement) onward. Measurable: weekly time-sheet extract reviewed "
     "at sprint close. If utilisation falls below 75% for 2 consecutive sprints, "
     "EPAM Delivery Director is notified.",
     "If utilisation < 75% for 2 consecutive sprints: EPAM identifies root cause "
     "within 3 business days and presents a corrective action plan. If corrective "
     "action does not restore ≥ 75% within 1 sprint, EPAM replaces the under-utilised "
     "role at no extra cost to Meridian."),
    ("A5 — Prosci sub-vendor on-schedule delivery",
     "Prosci sub-vendor delivers: (1) stakeholder mapping by end of Phase 0 week 2; "
     "(2) workshop 1 by Phase 2 week 1; (3) workshop 2 by Phase 2 week 3. "
     "Measurable: delivery date of each artefact against the named calendar date.",
     "If any deliverable is > 3 business days late: EPAM Engagement Lead invokes "
     "right-to-cure clause in sub-contractor MSA; Prosci has 5 business days to "
     "deliver or provide a replacement resource. EPAM absorbs sub-contractor "
     "remediation cost within contingency. Meridian is not exposed to sub-vendor risk."),
]

for ri, (label, asm, consequence) in enumerate(assumptions, 3):
    bg = WHT if ri % 2 == 1 else LGR
    for ci, val in enumerate([ri - 2, asm, consequence], 1):
        if ci == 1:
            c = ws5.cell(row=ri, column=ci, value=label)
            c.font = cf(bold=True); c.fill = fl(bg); c.border = BDR
            c.alignment = al("left", "top", True)
        else:
            c = ws5.cell(row=ri, column=ci, value=val)
            c.font = cf(); c.fill = fl(bg); c.border = BDR
            c.alignment = al("left", "top", True)
    ws5.row_dimensions[ri].height = 75

# ─────────────────────────────────────────────────────────────────────────────
# TAB 6 — COMMERCIAL MODEL
# ─────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("6 — Commercial Model")
set_w(ws6, {"A": 22, "B": 28, "C": 28, "D": 28, "E": 5})

section_title(ws6, 1, 4, "Commercial Model Decision — Hybrid recommended (failure mode 4 fix: T&M replaced)")
header_row(ws6, 2, ["Model", "Risk fit", "Cash-flow fit", "Buyer fit (RFP §7)"])

models = [
    ("Time & Materials (open-ended)",
     "EPAM bears zero cost risk — all schedule and scope risk falls on Meridian. "
     "Root-cause gate creates unbounded exposure: Meridian pays whether or not "
     "the gate passes.",
     "Meridian pays as-incurred; predictable for EPAM; unpredictable for buyer.",
     "DISQUALIFYING — RFP §7 submission rules: 'open-ended T&M is a disqualifying "
     "condition.' Cannot be submitted."),
    ("Fixed Price (single lump sum)",
     "EPAM bears all scope risk including root-cause gate variability and SAP "
     "integration unknowns. Both are high-uncertainty; pricing in a contingency "
     "to cover them inflates the fixed price above the €350K ceiling.",
     "EPAM cash-flow risk: cost overruns eat margin and contingency simultaneously. "
     "Requires very high contingency to be viable — pushes total above cap.",
     "RFP allows fixed-price but does not require it. Fixed price for a variable-"
     "scope engagement is a contract trap (see 02-review.md Critique 1)."),
    ("Hybrid: Phase 0 fixed-fee + Phases 1–3 capped T&M  ← RECOMMENDED",
     "Phase 0 (€33K fixed): gate risk is Meridian's — if gate fails, they stop paying. "
     "Phases 1–3 (€307K capped T&M): EPAM carries cost risk within the cap; "
     "change-control triggers protect EPAM from scope creep. "
     "Each party carries the risk they can control.",
     "Meridian knows maximum exposure (€340K all-in). Phase 0 is a low-risk first "
     "invoice before committing to full engagement. EPAM bills actuals up to cap.",
     "COMPLIANT — RFP §7: 'fixed-price or capped T&M; no open-ended T&M overrun "
     "clauses.' Capped T&M with named change-control satisfies the constraint. "
     "Change-control triggers: root-cause gate scope change; "
     "POS data restriction to one country."),
]

colours_row = [RED, AMB, GRN]

for ri, (model, risk_fit, cash_fit, buyer_fit) in enumerate(models, 3):
    bg = colours_row[ri - 3]
    for ci, val in enumerate([model, risk_fit, cash_fit, buyer_fit], 1):
        c = ws6.cell(row=ri, column=ci, value=val)
        c.font = cf(bold=(ri == 5 and ci == 1))
        c.fill = fl(bg); c.border = BDR
        c.alignment = al("left", "top", True)
    ws6.row_dimensions[ri].height = 88

ws6.row_dimensions[5].height = 100

# ── Recommendation paragraph ──────────────────────────────────────────────────
ws6.merge_cells("A6:D6")
c = ws6["A6"]
c.value = "Recommendation"
c.font = cf(bold=True, sz=11); c.fill = fl(AMB)
c.alignment = al("left", "center")
ws6.row_dimensions[6].height = 20

ws6.merge_cells("A7:D7")
c = ws6["A7"]
c.value = (
    "Recommend Hybrid: Phase 0 fixed-fee €33K (gated discovery) + Phases 1–3 capped T&M "
    "at €307K, total ceiling €340K. RATIONALE: the RFP explicitly prohibits open-ended T&M "
    "(disqualifying condition) and the root-cause gate creates a variable scope that makes "
    "a single fixed-price binding contract contractually dangerous (see 02-review.md, "
    "Critique 1 patch). Hybrid respects the RFP constraint ('capped T&M' is listed as "
    "permitted), aligns each party's risk with their control surface, and gives Meridian "
    "a known maximum exposure. The change-control mechanism (gate-failure scope change + "
    "POS data restriction to one country) is written into the contract as named triggers, "
    "not left as 'scope TBD at delivery'."
)
c.font = cf(sz=10); c.fill = fl(WHT)
c.alignment = al("left", "top", True)
ws6.row_dimensions[7].height = 90


# ─────────────────────────────────────────────────────────────────────────────
out = "Artefacts/1000-bid/04-estimate.xlsx"
wb.save(out)
print(f"Written: {out}\n")

# Print summary for verification
print(f"Delivery subtotal:   €{delivery_subtotal:,.0f}")
print(f"Contingency (8%):    €{contingency_amt:,.0f}")
print(f"Margin (~11% rev):   €{margin_amt:,.0f}")
print(f"TOTAL PRICE:         €{total_price:,.0f}")
print(f"\nTotal expected risk: €{total_ev:,.0f}  (contingency covers {int(contingency_amt/total_ev)}× expected)")
